import openpyxl


class CorpusCollection:
    def __init__(self, file_xlsx):
        self.data = openpyxl.load_workbook(file_xlsx)
        self.sheet = self.data.active

    def processing_corpus(self, label):
        # stating from row 2; column 3
        output_corpus = []
        label = ''
        # 28 the number is end of line
        for row in range(2, 28):
            sentence = self.sheet.cell(row=row, column=1).value
            if self.sheet.cell(row=row, column=3).value is not None:
                entity = self.sheet.cell(row=row, column=3).value
                label = 'position'
            elif self.sheet.cell(row=row, column=4).value is not None:
                entity = self.sheet.cell(row=row, column=4).value
                label = 'hobbies'
            elif self.sheet.cell(row=row, column=5).value is not None:
                entity = self.sheet.cell(row=row, column=5).value
                label = 'project'
            else:
                continue
            sentence_len = len(sentence)
            position_len = len(entity)
            for i in range(sentence_len - position_len + 1):
                if sentence[i:i + position_len] in entity:
                    print(i, i+position_len)
                    output_corpus.append((sentence, {'entities': [(i, i + position_len, label)]}))
        print(output_corpus)
        return output_corpus


if __name__ == '__main__':
    start = CorpusCollection('corpus.xlsx')
    TRAIN_DATA = start.processing_corpus('position')
